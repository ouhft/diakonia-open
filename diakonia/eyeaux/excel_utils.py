#!/usr/bin/python
# coding: utf-8
from openpyxl import Workbook, load_workbook

class NHSBTworkbook(Workbook):
    """
    Helper class to load, and query the NHSBT supplied Excel datasheet, extending the openpyxl Workbook model

    The mapping should look something like this:
    Col ID: Variable name - Description - NHIC Mapping

    01: RECIP_FORENAME - Recipient Forename - NHIC_TRA_1
    02: RECIP_SURNAME - Recipient Surname - NHIC_TRA_3
    03: RDOB - Recipient date of birth - NHIC_TRA_6
    04: RNHS_NO - Recipient NHS Number - NHIC_TRA_7
    05: RPOSTCODE - Recipient Postcode - NHIC_TRA_8
    06: recip_residence - Recipient country of residence - NHIC_TRA_13
    07: RSEX - Recipient gender - NHIC_TRA_14, NHIC_TRA_14_1
    08: RETHNIC - Recipient ethnicity - NHIC_TRA_15, NHIC_TRA_15_1
    09: RBG - Recipient blood group - NHIC_TRA_17, NHIC_TRA_17_1
    10: RRHESUS - Recipient rhesus blood group - NHIC_TRA_18, NHIC_TRA_18_1
    11: RECIP_ID - NHSBT recipient identification number - NHIC_TRA_19
    12: RDOD - recipient date of death - NHIC_TRA_21
    13: DONOR_FORENAME - donor forename - NHIC_TRA_22
    14: DONOR_SURNAME - donor surname - NHIC_TRA_24
    15: DDOB - donor date of birth - NHIC_TRA_27
    16: DNHS_NO - donor NHS Number - NHIC_TRA_28
    17: DONOR_POSTCODE - Donor postcode - NHIC_TRA_29
    18: DRESIDENCE - Donor counrty of residence - NHIC_TRA_34
    19: DSEX - Donor gender - NHIC_TRA_35, NHIC_TRA_35_1, NHIC_TRA_43, NHIC_TRA_43_1
    20: DETHNIC - Donor ethnicity - NHIC_TRA_36, NHIC_TRA_36_1, NHIC_TRA_45
    21: DBG - Donor blood group - NHIC_TRA_38, NHIC_TRA_48
    22: DRHESUS - Donor rhesus blood group - NHIC_TRA_39, NHIC_TRA_49
    23: DONOR_ID - NHSBT donor identification number - NHIC_TRA_40, NHIC_TRA_50
    24: DAGE - Donor age - NHIC_TRA_42
    25: DCOD - Donor cause of death - NHIC_TRA_44, NHIC_TRA_44_1
    26: DHTLV - Does the donor have HTLV - NHIC_TRA_46
    27: DSYPHILIS - Does the donor have syphilis - NHIC_TRA_46
    28: DPAST_ALCOHOL_ABUSE - past history of alcohol abuse in the donor - NHIC_TRA_46
    29: DPAST_CARDIO_DISEASE - past history of cardio disease in the donor - NHIC_TRA_46
    30: DPAST_DIABETES - past history of diabetes in the donor - NHIC_TRA_46
    31: DPAST_DRUG_ABUSE - past history of drug abuse in the donor - NHIC_TRA_46
    32: DPAST_HYPERTENSION - past history of hypertension in the donor - NHIC_TRA_46
    33: DPAST_LIVER_DISEASE - past history of liver disease in the donor - NHIC_TRA_46
    34: DPAST_OTHER - any other past history in the donor - NHIC_TRA_46
    35: DPAST_OTHER_DETAILS - details of other past histories - NHIC_TRA_46
    36: DPAST_SMOKER - past history of smoking in the donor - NHIC_TRA_46
    37: dpast_smoker_amount - past smoking amount - NHIC_TRA_46
    38: DPAST_TUMOUR - past history of tumour in the donor - NHIC_TRA_46
    39: DPAST_UTI - past history of UTI in the donor - NHIC_TRA_46
    40: DPAST_FAMILY_DIABETES - past history of diabetes in the donor family - NHIC_TRA_46
    41: DPAST_FAMILY_DIABETES_TYPE - what type of family diabetes - NHIC_TRA_46
    42: DIAL_AT_TX - was the recipient on dialysis at transplant - NHIC_TRA_55
    43: DIAL_AT_TX_TYPE - what type of dialysis was the recipient on - NHIC_TRA_56
    44: DIAL_START_DATE - what date did the recipient start dialysis - NHIC_TRA_57
    45: DIAL_END_DATE - what date did the recipient stop dialysis - NHIC_TRA_58
    46: NO_PREV_TX - had the recipient had any previous transplants - NHIC_TRA_59
    47: TX_DATE - date of recipients transplant - NHIC_TRA_66
    48: REC_UNIT - Transplant unit - NHIC_TRA_67
    49: RHOSP_NO - recipients hospital number - NHIC_TRA_68, NHIC_TRA_69
    50: DTYPE - donor type - NHIC_TRA_70, NHIC_TRA_70_1
    51: RLTNSHIP - relationship status of living donor transplants - NHIC_TRA_71, NHIC_TRA_71_1
    52: RELATIONSHIP_DETAILS - details of the relationship between recipient and donor - NHIC_TRA_71, NHIC_TRA_71_1
    53: LATERALITY - No description - NHIC_TRA_72
    54: TX_TYPE - Type of transplant performed - NHIC_TRA_73, NHIC_TRA_73_1
    55: LOCAL - was the transplant performed in the same centre as the donor - NHIC_TRA_74
    56: FAILDATE - what date did the transplant fail - NHIC_TRA_77
    57: AITX_TYPE - what type of antibody incompatible transplant was performed - NHIC_TRA_78, NHIC_TRA_81
    58: CIT_HRS - cold ischaemia time in hours - NHIC_TRA_83
    59: fWIT - functional warm ischaemia time in minutes - NHIC_TRA_84
    60: sWIT - standard warm ischaemia time in minutes - NHIC_TRA_84
    61: AMM - how many mismatches at the A locus - NHIC_TRA_85
    62: BMM - how many mismatches at the B locus - NHIC_TRA_85
    63: DRMM - how many mismatches at the DR locus - NHIC_TRA_85
    64: RECIP_HLA_SAMPLE_DATE - when was the last recipient HLA sample taken - NHIC_TRA_86
    65: Recip_First_A_Broad - Recipient HLA type - NHIC_TRA_88
    66: Recip_First_A_Split - Recipient HLA type - NHIC_TRA_88
    67: Recip_Second_A_Broad - Recipient HLA type - NHIC_TRA_88
    68: Recip_Second_A_Split - Recipient HLA type - NHIC_TRA_88
    69: Recip_First_B_Broad - Recipient HLA type - NHIC_TRA_88
    70: Recip_First_B_Split - Recipient HLA type - NHIC_TRA_88
    71: Recip_Second_B_Broad - Recipient HLA type - NHIC_TRA_88
    72: Recip_Second_B_Split - Recipient HLA type - NHIC_TRA_88
    73: Recip_First_C_Broad - Recipient HLA type - NHIC_TRA_88
    74: Recip_First_C_Split - Recipient HLA type - NHIC_TRA_88
    75: Recip_Second_C_Broad - Recipient HLA type - NHIC_TRA_88
    76: Recip_Second_C_Split - Recipient HLA type - NHIC_TRA_88
    77: Recip_First_DR_Broad - Recipient HLA type - NHIC_TRA_88
    78: Recip_First_DR_Split - Recipient HLA type - NHIC_TRA_88
    79: Recip_Second_DR_Broad - Recipient HLA type - NHIC_TRA_88
    80: Recip_Second_DR_Split - Recipient HLA type - NHIC_TRA_88
    81: Recip_First_DP_Broad - Recipient HLA type - NHIC_TRA_88
    82: Recip_First_DP_Split - Recipient HLA type - NHIC_TRA_88
    83: Recip_Second_DP_Broad - Recipient HLA type - NHIC_TRA_88
    84: Recip_Second_DP_Split - Recipient HLA type - NHIC_TRA_88
    85: Recip_First_DQ_Broad - Recipient HLA type - NHIC_TRA_88
    86: Recip_First_DQ_Split - Recipient HLA type - NHIC_TRA_88
    87: Recip_Second_DQ_Broad - Recipient HLA type - NHIC_TRA_88
    88: Recip_Second_DQ_Split - Recipient HLA type - NHIC_TRA_88
    89: DONOR_HLA_SAMPLE_DATE - When was the last donor HLA sample taken - ?
    90: Donor_First_A_Broad - Donor HLA type - ?
    91: Donor_First_A_Split - Donor HLA type - ?
    92: Donor_Second_A_Broad - Donor HLA type - ?
    93: Donor_Second_A_Split - Donor HLA type - ?
    94: Donor_First_B_Broad - Donor HLA type - ?
    95: Donor_First_B_Split - Donor HLA type - ?
    96: Donor_Second_B_Broad - Donor HLA type - ?
    97: Donor_Second_B_Split - Donor HLA type - ?
    98: Donor_First_C_Broad - Donor HLA type - ?
    99: Donor_First_C_Split - Donor HLA type - ?
    100: Donor_Second_C_Broad - Donor HLA type - ?
    101: Donor_Second_C_Split - Donor HLA type - ?
    102: Donor_First_DR_Broad - Donor HLA type - ?
    103: Donor_First_DR_Split - Donor HLA type - ?
    104: Donor_Second_DR_Broad - Donor HLA type - ?
    105: Donor_Second_DR_Split - Donor HLA type - ?
    106: Donor_First_DP_Broad - Donor HLA type - ?
    107: Donor_First_DP_Split - Donor HLA type - ?
    108: Donor_Second_DP_Broad - Donor HLA type - ?
    109: Donor_Second_DP_Split - Donor HLA type - ?
    110: Donor_First_DQ_Broad - Donor HLA type - ?
    111: Donor_First_DQ_Split - Donor HLA type - ?
    112: Donor_Second_DQ_Broad - Donor HLA type - ?
    113: Donor_Second_DQ_Split - Donor HLA type - ?
    114: CYCLOSPORIN - Immunosuppresion and induction therapy - NHIC_TRA_91
    115: MYCOPHENOLATE_MOFETIL - Immunosuppresion and induction therapy - NHIC_TRA_91
    116: OTHER_IMMUNO_DRUG - Immunosuppresion and induction therapy - NHIC_TRA_91
    117: IMMUNO_TEXT - Immunosuppresion and induction therapy - NHIC_TRA_91
    118: PREDNISOLONE_PREDNISONE - Immunosuppresion and induction therapy - NHIC_TRA_91
    119: TACROLIMUS - Immunosuppresion and induction therapy - NHIC_TRA_91
    120: AZATHIOPRINE - Immunosuppresion and induction therapy - NHIC_TRA_91
    121: ALG_ATG - Immunosuppresion and induction therapy - NHIC_TRA_91
    122: OKT3 - Immunosuppresion and induction therapy - NHIC_TRA_91
    123: RCMV - Recipient CMV status - NHIC_TRA_97
    124: RHIV - Recipient HIV status - NHIC_TRA_105
    125: DEBV - Donor EBV status - NHIC_TRA_109
    126: DTOXO - Donor TOXO status - NHIC_TRA_113
    127: DHIV - Donor HIV status - NHIC_TRA_115
    128: DGF - Did the recipient have delayed graft function - NHIC_TRA_124
    129: RESUME_DIALYSIS_DATE - when did the recipient resume dialysis post transplant - NHIC_TRA_126
    130: RHEIGHT - recipient height - NHIC_TRA_127
    131: RWEIGHT - recipient weight - NHIC_TRA_130
    132: serum3 - serum creatinine at 3 months - NHIC_TRA_142
    133: scdate3 - date of serum creatinine at 3 months - NHIC_TRA_144
    134: serum12 - serum creatinine at 12 months - NHIC_TRA_142
    135: scdate12 - date of serum creatinine at 12 months - NHIC_TRA_144
    136: serum24 - serum creatinine at 24 months - NHIC_TRA_142
    137: scdate24 - date of serum creatinine at 24 months - NHIC_TRA_144
    138: serum36 - serum creatinine at 36 months - NHIC_TRA_142
    139: scdate36 - date of serum creatinine at 36 months - NHIC_TRA_144
    140: serum48 - serum creatinine at 48 months - NHIC_TRA_142
    141: scdate48 - date of serum creatinine at 48 months - NHIC_TRA_144
    142: serum60 - serum creatinine at 60 months - NHIC_TRA_142
    143: scdate60 - date of serum creatinine at 60 months - NHIC_TRA_144
    144: RBHV - recipient BHV status - NHIC_TRA_230
    145: RHCV - Recipient HCV status - NHIC_TRA_232
    146: DHBCAB - donor Hep B core antigen status - NHIC_TRA_233
    147: DHBSAG - donor hep B surface antigen status - NHIC_TRA_234
    148: DHCV - Donor HCV status - NHIC_TRA_235
    149: eGFR3 - recipient eGFR at 3 months - NHIC_TRA_236
    150: eGFR12 - recipient eGFR at 12 months - NHIC_TRA_236
    151: eGFR24 - recipient eGFR at 24 months - NHIC_TRA_236
    152: eGFR36 - recipient eGFR at 36 months - NHIC_TRA_236
    153: eGFR48 - recipient eGFR at 48 months - NHIC_TRA_236
    154: eGFR60 - recipient eGFR at 60 months - NHIC_TRA_236
    XXX: DCMV - Donor CMV - NHIC_TRA_107 .. NB: missing
    """

    __workbook__ = None
    __current_row_id__ = 2  # First row is headers, so start with data
    __current_row_data__ = dict()  # Store the row in a dictionary based on the column titles
    __headers_by_id__ = dict()
    __headers_by_title__ = dict()

    def __init__(self, workbook=None):
        if workbook is not None:
            self.__workbook__ = workbook
            self.__load_header_dictionaries__()
        super(NHSBTworkbook, self).__init__()

    def __load_header_dictionaries__(self):
        # Find the headers in row 1 of this worksheet, and store in the first dictionary, and flip for the second
        for column_id in range(1, self.worksheet.max_column + 1):
            self.__headers_by_id__[column_id] = self.worksheet.cell(row=1, column=column_id).value
            self.__headers_by_title__[self.worksheet.cell(row=1, column=column_id).value] = column_id

    def load_xlsx(self, filename=None):
        if isinstance(filename, str):
            self.__workbook__ = load_workbook(filename)
            self.__load_header_dictionaries__()
            return True
        return False

    def load_row(self, row_id=0):
        """
        Load the specified row contents into a dictionary keyed on the column title. The title is
        reduced to lower case so that it matches model naming conventions and is consistent.

        :param row_id: int representing the worksheet row. Must be 2 or greater as row 1 is titles
        :return: dict() of row data if row id provided, otherwise, int of the current row id
        """
        if row_id > 1:
            self.__current_row_id__ = row_id
            for column_id in range(1, self.worksheet.max_column + 1):
                self.__current_row_data__[self.__headers_by_id__[column_id].lower()] =\
                    self.worksheet.cell(row=row_id, column=column_id).value
            return self.__current_row_data__
        return self.__current_row_id__

    def headers(self, key=None, value=None):
        if key is not None:
            return self.__headers_by_id__[key]
        if value is not None:
            return self.__headers_by_title__[value]
        return self.__headers_by_id__

    @property
    def workbook(self):
        return self.__workbook__

    @property
    def worksheet(self):
        return self.workbook.active


# Clean up utilities that come in handy all over the place
def int_as_str(value):
    """
    Convert file values that appear as integers, into string results that look like they do in the application (i.e. 1234.0 in file -> 1234 as output)
    :param value:
    :return: string
    """
    output = ""
    try:
        output = int(value)
    except ValueError:
        try:
            output = int(float(value))
        except ValueError:  # "could not convert string to float"
            pass
    except TypeError:  # For when the value is None
        pass
    return str(output)


class ReadOnlyWorkbook(object):
    __workbook__ = None
    __worksheet__ = None
    __max_columns__ = 0
    __headers_by_id__ = dict()
    __headers_by_title__ = dict()
    __current_row_id__ = 2  # First row is headers, so start with data
    __current_row_data__ = dict()  # Store the row in a dictionary based on the column titles

    def __init__(self, filename=""):
        if filename is not "":
            self.load_workbook(filename=filename)

    def load_workbook(self, filename="", column_count=1):
        if isinstance(filename, str) and filename is not "":
            self.__workbook__ = load_workbook(filename=filename, read_only=True)
            self.__worksheet__ = self.__workbook__.active
            self.__max_columns__ = column_count + 1  # NB: Excel references start at 1, not 0
            print("DEBUG: Max columns = {0}".format(self.__max_columns__))
            for column_id in range(1, self.__max_columns__):
                self.__headers_by_id__[column_id] = self.__worksheet__.cell(row=1, column=column_id).value
                self.__headers_by_title__[self.__worksheet__.cell(row=1, column=column_id).value] = column_id

            print("DEBUG: load_workbook: headers = {0}".format(self.__headers_by_id__))
            return True
        return False

    def get_rows(self):
        if self.__worksheet__ is not None:
            return self.__worksheet__.iter_rows()
        return None

    def load_row(self, row=None):
        """
        Load the current row contents into a dictionary keyed on the column title. The title is
        reduced to lower case so that it matches model naming conventions and is consistent.

        :param row: Worksheet row to be loaded into dict
        :return: dict() of row data if row id provided, otherwise, int of the current row id
        """
        column_id = 1
        for cell in row:
            self.__current_row_data__[self.__headers_by_id__[column_id].lower()] = cell.value
            column_id += 1
        return self.__current_row_data__



